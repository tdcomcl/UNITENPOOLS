"""
Script para importar datos del Excel a la base de datos SQLite.
"""
import openpyxl
from database import Database
import sys


def importar_desde_excel(excel_path: str, db_path: str = "piscinas.db"):
    """Importa los datos del Excel a la base de datos."""
    db = Database(db_path)
    
    print(f"Leyendo archivo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Leer encabezados
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            headers[val] = col
    
    print(f"Columnas encontradas: {list(headers.keys())}")
    
    # Procesar filas
    clientes_importados = 0
    responsables_creados = {}
    
    # Primero, crear responsables únicos
    print("\nIdentificando responsables...")
    responsables_set = set()
    for row in range(2, ws.max_row + 1):
        responsable_val = ws.cell(row, headers.get('Responsable', 0)).value
        if responsable_val and responsable_val not in ['Responsable', None]:
            responsables_set.add(str(responsable_val).strip())
    
    print(f"Responsables encontrados: {list(responsables_set)}")
    
    # Crear responsables en la base de datos
    for resp_name in responsables_set:
        resp_id = db.agregar_responsable(resp_name)
        responsables_creados[resp_name] = resp_id
    
    # Importar clientes
    print("\nImportando clientes...")
    for row in range(2, ws.max_row + 1):
        # Obtener datos del cliente
        nombre = ws.cell(row, headers.get('Nombre cliente', 0)).value
        if not nombre or nombre == 'Nombre cliente':
            continue
        
        direccion = ws.cell(row, headers.get('Dirección', 0)).value
        comuna = ws.cell(row, headers.get('Comuna', 0)).value
        celular = ws.cell(row, headers.get('Celular', 0)).value
        responsable_val = ws.cell(row, headers.get('Responsable', 0)).value
        dia_atencion = ws.cell(row, headers.get('día de atención', 0)).value
        precio_val = ws.cell(row, headers.get('precio', 0)).value
        
        # Limpiar y procesar datos
        nombre = str(nombre).strip()
        direccion = str(direccion).strip() if direccion else None
        comuna = str(comuna).strip() if comuna else None
        celular = str(celular).strip() if celular else None
        dia_atencion = str(dia_atencion).strip() if dia_atencion else None
        
        # Obtener responsable_id
        responsable_id = None
        if responsable_val and responsable_val not in ['Responsable', None]:
            responsable_name = str(responsable_val).strip()
            if responsable_name in responsables_creados:
                responsable_id = responsables_creados[responsable_name]
        
        # Procesar precio
        precio_por_visita = 0
        if precio_val:
            try:
                # Intentar convertir a número
                if isinstance(precio_val, (int, float)):
                    precio_por_visita = float(precio_val)
                else:
                    precio_str = str(precio_val).replace('$', '').replace(',', '').strip()
                    precio_por_visita = float(precio_str)
            except (ValueError, TypeError):
                precio_por_visita = 0
        
        # Insertar cliente
        try:
            cliente_id = db.agregar_cliente(
                nombre=nombre,
                direccion=direccion,
                comuna=comuna,
                celular=celular,
                responsable_id=responsable_id,
                dia_atencion=dia_atencion,
                precio_por_visita=precio_por_visita
            )
            clientes_importados += 1
            if clientes_importados % 50 == 0:
                print(f"  Importados {clientes_importados} clientes...")
        except Exception as e:
            print(f"  Error al importar cliente {nombre}: {e}")
            continue
    
    print(f"\n✓ Importación completada!")
    print(f"  - Responsables creados: {len(responsables_creados)}")
    print(f"  - Clientees importados: {clientes_importados}")
    
    return clientes_importados


if __name__ == "__main__":
    excel_path = "Base de Datos United al 28 oct 2025.xlsx"
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    try:
        importar_desde_excel(excel_path)
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {excel_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error durante la importación: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

